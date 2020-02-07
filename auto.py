from THS.THSTrader import THSTrader
import tushare as ts
import talib as ta

import openpyxl

import time
import math

# 读写Excel的方法----------------------------------------


def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
    workbook.save(path)
    print("xlsx格式表格写入数据成功！")


def read_excel_xlsx(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    l = []

    for row in sheet.rows:
        tmp = []
        for cell in row:
            tmp.append(cell.value)
        l.append(tmp.copy())
    return l

# 获取必要数据的方法----------------------------------------


def get_boll(stock_code: str, ma_day: int):
    ma_type = 'ma'+str(ma_day)

    pre_data = ts.get_k_data(stock_code)
    close_list = pre_data.close.values[-100:]
    price = get_now(stock_code)
    close_list.append(price)

    up, mid, dn = ta.BBANDS(close_list, timeperiod=ma_day,
                            nbdevup=2, nbdevdn=2, matype=0)
    return round(up[-1], 3), round(mid[-1], 2), round(dn[-1], 2)


def get_rsi(stock_code: str, rsi_day: int):
    pre_data = ts.get_k_data(stock_code)
    close_list = pre_data.close.values[-100:]

    rsi = ta.RSI(close_list, rsi_day)
    return rsi[-1]


def get_price(stock_code: str) -> float:
    return round(float(ts.get_realtime_quotes(stock_code).price.values[0]), 2)


def get_position() -> list:
    position = trader.get_position()
    return position


# 脚本过程----------------------------------------


# 存放交易记录的文件
book_name_xlsx = 'log.xlsx'
sheet_name_xlsx = 'sheet1'
write_list = []
# 新建/追加记录
try:
    write_list = read_excel_xlsx(book_name_xlsx, sheet_name_xlsx)
except:
    write_list.append(['时间', '价格', '成交数量', '成交金额'])

# 同花顺客户端过程
trader = THSTrader(r"C:\\同花顺软件\\同花顺\\xiadan.exe")    # 连接客户端
position = get_position()  # 持仓信息
fortune = trader.get_balance()['可用金额'] #可支配金额，每次交易前都要重新获取一次
spy_code_list = {}  # 被清仓的代码与清仓价值, {code:value}
for item in position:
    fortune = trader.get_balance()['可用金额']
    code = (str(item['证券代码']).zfill(6))
    cost = item['成本价']
    price = item['市价']
    value = item['市值']  # 现在的价值
    rsi = get_rsi(code, 6)
    if rsi > 97:
        if price > cost:  # 盈利
            try:
                trader.sell(stock_no=code,
                            amount=item['可用余额'],
                            price=price)  # 清仓，记录清仓价值
                spy_code_list[code] = value  # 添加到观察的列表中
                print("{}交易成功，已清仓。".format(code))
                write_list.append([
                    time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()),
                    price,
                    item['可用余额'],
                    item['可用余额']*price
                ])
                print(write_list)
                write_excel_xlsx(book_name_xlsx, sheet_name_xlsx, write_list)
            except:
                print("{}清仓失败！".format(code))
        else:
            value2 = math.sqrt(cost)*value/math.sqrt(price)  # 交易后的价值
            pay = (value2-value)/price  # 正买负卖
            try:
                flag = 1
                if pay > 0:  # 交易
                    if pay//100*100*price > fortune:
                        amount = fortune//price//100*100
                    else:
                        amount = pay//100*100
                    if amount == 0:
                        continue
                    trader.buy(stock_no=code,
                               amount=amount,
                               price=price)
                    flag = -1
                else:
                    if -pay//100*100*price > value:
                        amount = int(value//price)
                    else:
                        amount = -pay//100*100
                    if amount == 0:
                        continue
                    trader.sell(stock_no=code,
                                amount=amount,
                                price=price)
                write_list.append([
                    time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()),
                    price,
                    flag*amount,
                    flag*amount*price
                ])
                write_excel_xlsx(book_name_xlsx, sheet_name_xlsx, write_list)
                print("{}交易成功。".format(code))

            except:
                print("{}交易失败！".format(code))
for code in spy_code_list.keys():
    fortune = trader.get_balance()['可用金额']
    rsi = get_rsi(code, 6)
    price = get_price(code)
    if rsi < 3:  # 有过清仓行为,买入value*0.95
        try:
            if (spy_code_list[code] * 0.95//price)*price > fortune:
                amount = fortune//price//100*100
            else:
                amount = (spy_code_list[code] * 0.95//price)
            if amount == 0:
                continue
            trader.buy(stock_no=code,
                       amount=amount,
                       price=price)
            del spy_code_list[code]
            write_list.append([
                time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()),
                price,
                -amount,
                -amount*price
            ])
            write_excel_xlsx(book_name_xlsx, sheet_name_xlsx, write_list)
            print("{}交易成功。".format(code))
        except:
            print("{}交易失败！".format(code))
    else:
        pass  # 无操作

# 时间，价格，成交数量，成交金额
